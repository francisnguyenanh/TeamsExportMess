"""
3_export.py
===========
Đọc config.json → export từng channel ra file Excel riêng.

Cách dùng:
    python 3_export.py
    python 3_export.py --config my_config.json
"""

import argparse
import html
import json
import re
import requests
import openpyxl
from pathlib import Path
from datetime import datetime, timezone
from openpyxl.styles import Font, PatternFill, Alignment

from get_token_helper import get_valid_token
from get_token import save_token       # để lưu token mới nếu cần refresh

CONFIG_FILE = "config.json"


# ── Helpers ───────────────────────────────────────────────────────────────────

def make_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def strip_html(text: str) -> str:
    """Xóa HTML tags, decode HTML entities, chuẩn hoá khoảng trắng."""
    # 1. Block-level tags → newline
    text = re.sub(r"<br\s*/?>\s*", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p>\s*",       "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p[^>]*>",      "",   text, flags=re.IGNORECASE)
    text = re.sub(r"<div[^>]*>",    "",   text, flags=re.IGNORECASE)
    text = re.sub(r"</div>\s*",     "\n", text, flags=re.IGNORECASE)
    # 2. Strip all remaining tags
    text = re.sub(r"<[^>]+>", "", text)
    # 3. Decode HTML entities (&nbsp; → space, &amp; → &, etc.)
    text = html.unescape(text)
    # 4. Replace non-breaking spaces and other Unicode spaces with regular space
    text = text.replace("\u00a0", " ")   # &nbsp;
    text = text.replace("\u200b", "")    # zero-width space
    text = text.replace("\u200c", "")    # zero-width non-joiner
    text = text.replace("\ufeff", "")    # BOM
    # 5. Collapse lines: trim each line, drop lines that are only whitespace
    lines = [line.strip() for line in text.splitlines()]
    # Collapse runs of more than 2 consecutive blank lines to a single blank line
    cleaned = []
    blank_count = 0
    for line in lines:
        if line == "":
            blank_count += 1
            if blank_count <= 1:
                cleaned.append(line)
        else:
            blank_count = 0
            cleaned.append(line)
    return "\n".join(cleaned).strip()


def parse_dt(iso_str: str) -> str:
    """Chuyển ISO datetime → chuỗi đọc được, múi giờ UTC+7."""
    if not iso_str:
        return ""
    dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
    # Chuyển sang UTC+7 (Việt Nam)
    from datetime import timedelta
    dt_vn = dt.astimezone(timezone(timedelta(hours=7)))
    return dt_vn.strftime("%Y-%m-%d %H:%M:%S")


def parse_msg(msg: dict, is_reply: bool = False) -> dict:
    """Trích xuất thông tin cần thiết từ message object."""
    sender_obj = msg.get("from") or {}
    user_obj   = sender_obj.get("user") or sender_obj.get("application") or {}
    content    = strip_html(msg.get("body", {}).get("content", ""))

    return {
        "type":       "↩ Reply" if is_reply else "💬 Message",
        "datetime":   parse_dt(msg.get("createdDateTime", "")),
        "sender":     user_obj.get("displayName", "Unknown"),
        "content":    content,
        "message_id": msg.get("id", ""),
    }


# ── Graph API ─────────────────────────────────────────────────────────────────

def api_get(url: str, headers: dict) -> dict:
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code == 401:
        raise PermissionError("TOKEN_EXPIRED")
    if resp.status_code == 403:
        raise PermissionError(f"ACCESS_DENIED: {url}")
    resp.raise_for_status()
    return resp.json()


def fetch_messages(team_id: str, channel_id: str, headers: dict,
                   date_from: str = "", date_to: str = "",
                   include_replies: bool = True) -> list[dict]:
    """
    Lấy toàn bộ tin nhắn trong channel.
    Tự động xử lý phân trang qua @odata.nextLink.
    """
    url = (f"https://graph.microsoft.com/v1.0"
           f"/teams/{team_id}/channels/{channel_id}/messages")
    rows = []
    page_num = 0

    while url:
        page_num += 1
        data = api_get(url, headers)

        for msg in data.get("value", []):
            if msg.get("messageType") != "message":
                continue

            # Lọc theo ngày (so sánh chuỗi ISO — hoạt động vì định dạng cố định)
            created = msg.get("createdDateTime", "")
            if date_from and created < date_from:
                continue
            if date_to and created > date_to + "T23:59:59Z":
                continue

            rows.append(parse_msg(msg, is_reply=False))

            # Lấy replies của thread này
            if include_replies:
                replies_url = (f"https://graph.microsoft.com/v1.0"
                               f"/teams/{team_id}/channels/{channel_id}"
                               f"/messages/{msg['id']}/replies")
                try:
                    rdata = api_get(replies_url, headers)
                    for r in sorted(rdata.get("value", []),
                                    key=lambda x: x.get("createdDateTime", "")):
                        if r.get("messageType") == "message":
                            rows.append(parse_msg(r, is_reply=True))
                except PermissionError:
                    raise   # Cho ra ngoài xử lý
                except Exception:
                    pass    # Bỏ qua lỗi nhỏ ở reply

        next_url = data.get("@odata.nextLink")
        url = next_url
        print(f"    Trang {page_num} — đã xử lý {len(rows)} hàng...", end="\r")

    print()  # Xuống dòng sau khi xong
    return rows


# ── Excel ─────────────────────────────────────────────────────────────────────

HEADER_BG  = "2F5496"
REPLY_BG   = "F5F5F5"
HEADER_COLS = ["Type", "DateTime (UTC+7)", "Sender", "Content", "MessageID"]
COL_WIDTHS  = {"A": 12, "B": 22, "C": 22, "D": 80, "E": 36}


def write_excel(rows: list[dict], output_path: Path, sheet_name: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header row
    ws.append(HEADER_COLS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows
    for row in rows:
        ws.append([
            row["type"], row["datetime"], row["sender"],
            row["content"], row["message_id"],
        ])
        row_num = ws.max_row
        # Tô nền nhạt cho Reply để phân biệt
        if row["type"].startswith("↩"):
            for cell in ws[row_num]:
                cell.fill = PatternFill("solid", fgColor=REPLY_BG)

        # Wrap text cột Content
        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True)

    # Độ rộng cột
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main(config_path: str = CONFIG_FILE):
    config     = json.loads(Path(config_path).read_text(encoding="utf-8"))
    channels   = config.get("channels", [])
    exp_cfg    = config.get("export", {})
    date_from  = exp_cfg.get("date_from", "")
    date_to    = exp_cfg.get("date_to", "")
    output_dir = Path(exp_cfg.get("output_dir", "output"))
    inc_reply  = exp_cfg.get("include_replies", True)
    browser    = config.get("browser", "edge")

    if not channels:
        print("⚠️  Không có channel nào trong config.json!")
        return

    # Lấy token hợp lệ (tự làm mới nếu cần)
    token   = get_valid_token(browser=browser)
    headers = make_headers(token)

    total   = len(channels)
    success = 0
    skipped = 0

    print(f"\n📦 Bắt đầu export {total} channel(s)...\n")

    for idx, ch in enumerate(channels, 1):
        team_name    = ch.get("team_name", "unknown_team")
        channel_name = ch.get("channel_name", "unknown_channel")
        team_id      = ch["team_id"]
        channel_id   = ch["channel_id"]

        print(f"[{idx}/{total}] 📂 {team_name} / #{channel_name}")

        rows = None
        for attempt in range(2):   # Thử tối đa 2 lần (1 lần nếu token hết giữa chừng)
            try:
                rows = fetch_messages(
                    team_id, channel_id, headers,
                    date_from=date_from,
                    date_to=date_to,
                    include_replies=inc_reply,
                )
                break   # Thành công → thoát vòng retry

            except PermissionError as e:
                err = str(e)
                if "TOKEN_EXPIRED" in err and attempt == 0:
                    # Token hết hạn giữa chừng → làm mới và thử lại
                    print("  🔄 Token hết hạn, đang làm mới...")
                    from get_token import get_token_sso
                    token = get_token_sso(browser=browser)
                    if token:
                        save_token(token)
                        headers = make_headers(token)
                    else:
                        print("  ❌ Không làm mới được token!")
                        break
                else:
                    print(f"  ⛔ Bỏ qua — {err}")
                    skipped += 1
                    break

        if rows is None:
            continue

        if not rows:
            print("  ⚠️  Không có tin nhắn nào (hoặc ngoài khoảng ngày lọc).\n")
            continue

        # Tạo tên file an toàn
        safe = re.sub(r'[\\/*?:"<>|]', "_", f"{team_name}_{channel_name}")
        today = datetime.now().strftime("%Y-%m")
        output_path = output_dir / f"{safe}_{today}.xlsx"

        write_excel(rows, output_path, sheet_name=channel_name)
        print(f"  ✅ {len(rows)} hàng → {output_path}\n")
        success += 1

    print("─" * 50)
    print(f"🎉 Hoàn thành!  Thành công: {success}  |  Bỏ qua: {skipped}  |  Tổng: {total}")
    print(f"📁 File Excel nằm trong thư mục: {output_dir.resolve()}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=CONFIG_FILE,
                        help=f"Đường dẫn file config (mặc định: {CONFIG_FILE})")
    args = parser.parse_args()
    main(config_path=args.config)
